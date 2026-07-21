import openpyxl
from openpyxl.utils import range_boundaries


def _merged_fill_map(sheet):
    fill = {}
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        top_left_value = sheet.cell(row=min_row, column=min_col).value
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                fill[(row, col)] = top_left_value
    return fill


def _get_unique_sheet_name(wb, name):
    name = name[:31]
    if name not in wb.sheetnames:
        return name
    
    for i in range(1, 100):
        suffix = f"_{i}"
        new_name = f"{name[:31-len(suffix)]}{suffix}"
        if new_name not in wb.sheetnames:
            return new_name
    return name


def extract_all_data(file_path, output_path):
    """
    Read every worksheet/row/column of file_path and write a clean,
    macro-free .xlsx copy at output_path.
    """
    src_wb = openpyxl.load_workbook(file_path, data_only=True, keep_vba=False)
    out_wb = openpyxl.Workbook()
    # Remove default sheet created by openpyxl
    if "Sheet" in out_wb.sheetnames:
        out_wb.remove(out_wb["Sheet"])

    report = {}
    for sheet in src_wb.worksheets:
        fill_map = _merged_fill_map(sheet)

        
        unique_name = _get_unique_sheet_name(out_wb, sheet.title)
        out_ws = out_wb.create_sheet(title=unique_name)

        max_row = sheet.max_row or 0
        max_col = sheet.max_column or 0
        non_empty = 0

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                src_cell = sheet.cell(row=row, column=col)
                value = fill_map.get((row, col), src_cell.value)
                
                if value is not None:
                    non_empty += 1
                    
                    new_cell = out_ws.cell(row=row, column=col, value=value)
                    
                    new_cell.number_format = src_cell.number_format

        report[sheet.title] = {
            "rows": max_row,
            "cols": max_col,
            "non_empty_cells": non_empty,
            "merged_ranges_flattened": len(sheet.merged_cells.ranges),
        }

    out_wb.save(output_path)
    src_wb.close()
    return report


def export_sheets_to_csv(clean_xlsx_path, out_dir):
    """Split a clean workbook into one CSV per sheet."""
    import csv
    import os

    os.makedirs(out_dir, exist_ok=True)
    wb = openpyxl.load_workbook(clean_xlsx_path, data_only=True)
    written = []
    for sheet in wb.worksheets:
        safe_name = "".join(c if c.isalnum() else "_" for c in sheet.title)
        csv_path = os.path.join(out_dir, f"{safe_name}.csv")
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for row in sheet.iter_rows(values_only=True):
                writer.writerow(["" if v is None else v for v in row])
        written.append(csv_path)
    wb.close()
    return written
